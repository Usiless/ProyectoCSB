from itertools import islice
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
import numpy as np

wb = load_workbook('p1.xlsx')
ws = wb.active
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'Prueba'
peso_evi = 0.5
peso_inst = 0.2
peso_virt = 0.2
peso_rsa = 0.1

data = ws.values
print('datos: ', data)
cols = next(data)[1:]
print('columnas: ', cols)
data = list(data)
print('datos: ', data)
idx = [r[0] for r in data]
print('indices: ', idx)
data = (islice(r, 1, None) for r in data)
print('datos: ', data)
df = DataFrame(data, index=idx, columns=cols)
print('dataframe: ', df)
df.dropna(inplace=True)
print('dataframe: ', df)


def extrae_valores(fila, a, b):
    pl = int(fila[b])
    tp = int(fila[a])
    porc = 0
    porc_tot = 0
    porc = pl / tp * 100
    if a/2==1:
        porc_tot = porc * peso_evi
    elif a/2==2:
        porc_tot = porc * peso_inst
    elif a/2==3:
        porc_tot = porc * peso_virt
    elif a/2==4:
        porc_tot = porc * peso_rsa
    return porc, porc_tot


def guardar(wb2, ws2, df):
    for r in dataframe_to_rows(df, index=True, header=True):
        ws2.append(r)
    print(list(ws2.values))
    for cell in ws2['A'] + ws2[1]:
        cell.style = 'Pandas'
    wb2.save('p2.xlsx')


def procesos_sin_formula(df):
    db = df.copy()
    porc_sum = np.array([])
    for t in range(2,10,2):
        porc = np.array([])
        porc_tot = np.array([])
        for fila in df.values:
            aux1, aux2 = extrae_valores(fila, t, t+1)
            porc = np.append(porc, aux1)
            porc_tot = np.append(porc_tot, aux2)
        a=t*2
        db.insert(a, "Cálculo", porc, allow_duplicates=True)
        db.insert(a+1, "Porcentaje", porc_tot, allow_duplicates=True)
    for r in dataframe_to_rows(db, index=False, header=False):
        aux1 = r[5]+r[9]+r[13]+r[17]
        porc_sum = np.append(porc_sum, aux1)
    idx = len([r for r in db])
    db.insert(idx, "Porcentaje total", porc_sum)
    return db


df = procesos_sin_formula(df)
guardar(wb2, ws2, df)
