from decimal import ROUND_DOWN
from itertools import islice
from math import trunc
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from pandas import DataFrame
import numpy as np

def planilla_proc_detallado(cabecera,tareas,detalle,cols):
    wb = load_workbook('Planilla_Tareas_Materia.xlsx')
    prof = cabecera[1]
    grado = cabecera[0]
    asig = cabecera[2]
    ws = wb.active
    
    cell = ws.cell(row=6, column=1)  
    cell.value = 'Profesor/a: ' + prof

    cell = ws.cell(row=6, column=7)  
    cell.value = 'Grado: ' + grado

    cell = ws.cell(row=5, column=12)
    if cabecera[3]==1:
        cell.value = "Etapa: Primera"
    else:
        cell.value = "Etapa: Segunda"

    cell = ws.cell(row=6, column=17)  
    cell.value = 'Asignatura: ' + asig

    cell = ws.cell(row=5, column=23)  
    cell.value = 'Año: ' + str(cabecera[4])

    for index, row in enumerate(tareas):
        if index !=0 :
            if row[4]==1:
                acum_ini += tareas[index-1][4]
                acum_fin += row[4]
            else: 
                if tareas[index-1][4]!=1:
                    acum_ini += tareas[index-1][4]
                    acum_fin += row[4]
                else: 
                    acum_ini += tareas[index-1][4]
                    acum_fin += row[4]
        else:
            if row[4]==1:
                acum_ini = 3
                acum_fin = 3 
            else:
                acum_ini = 3
                acum_fin = 3 + row[4]-1

        ws.merge_cells(start_row=8,start_column=acum_ini,end_row=8,end_column=acum_fin)
        cell = ws.cell(row=8, column=acum_ini)  
        cell.value = 'Tarea N°: ' + str(row[0])
        cell.alignment = Alignment(horizontal='center', vertical='center')  

        ws.merge_cells(start_row=9,start_column=acum_ini,end_row=9,end_column=acum_fin)
        cell = ws.cell(row=9, column=acum_ini)  
        if row[3] is None:
            cell.value = 'Capacidad: ' 
        else: 
            cell.value = 'Capacidad: ' + str(row[3])
        cell.alignment = Alignment(horizontal='left', vertical='center')  

        
        ws.merge_cells(start_row=10,start_column=acum_ini,end_row=10,end_column=acum_fin)
        cell = ws.cell(row=10, column=acum_ini)  
        cell.value = 'Tema: ' + row[1]
        cell.alignment = Alignment(horizontal='left', vertical='center')  

        ws.merge_cells(start_row=11,start_column=acum_ini,end_row=11,end_column=acum_fin)
        cell = ws.cell(row=11, column=acum_ini)  
        cell.value = 'Fecha: ' + str(row[2])
        cell.alignment = Alignment(horizontal='left', vertical='center')  

    idx = [i+1 for i in range(0,len(detalle))]
    data = detalle
    df = DataFrame(data, index=idx, columns=cols)
    df.dropna(inplace=True)

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    for rowy,row in enumerate(dataframe_to_rows(df, index=True, header=True), start=12):
        for colx, value in enumerate(row, start=1):
            colletfin = get_column_letter(colx)
            ws.cell(column=colx, row=rowy, value=value) 
            ws.cell(column=colx, row=rowy).border = thin_border
            if colx > 2:
                ws.cell(column=colx, row=13, value=1)
            if rowy > 12:
                if colx > 3:
                    if colx < 24:
                        ws.cell(column=24, row=rowy, value=f'=SUM(C{rowy}:{colletfin}{rowy})')

    wb.save(f'planillas/Planilla_{asig}.xlsx')

def planilla_proc_grado(cabecera,detalle,cols):
    wb = load_workbook('Planilla_Tareas_Grado.xlsx')
    prof = cabecera[1]
    grado = cabecera[0]
    asig = cabecera[2]
    ws = wb.active
    
    cell = ws.cell(row=6, column=1)  
    cell.value = 'Profesor/a: ' + prof

    cell = ws.cell(row=6, column=11)  
    cell.value = 'EEB: ' + cabecera[5]

    cell = ws.cell(row=6, column=19)  
    cell.value = 'Grado: ' + grado

    cell = ws.cell(row=5, column=10)
    if cabecera[3]==1:
        cell.value = "Periodo de Abril a Junio año Lectivo " + str(cabecera[4])
    else:
        cell.value = "Periodo de Julio a Noviembre año Lectivo " + str(cabecera[4])

    for index, row in enumerate(cols):
        if index !=0 :
            acum_ini += 3
        else:
            acum_ini = 3

        cell = ws.cell(row=8, column=acum_ini)  
        cell.value = row[0]
        cell = ws.cell(row=9, column=acum_ini)  
        cell.value = "TP"
        cell = ws.cell(row=9, column=acum_ini+1)  
        cell.value = "PL"
        cell = ws.cell(row=9, column=acum_ini+2)  
        cell.value = "Calf. Final"
        cell.alignment = Alignment(horizontal='center', vertical='center')

    idx = [i+1 for i in range(0,len(detalle))]
    data = detalle
    df = DataFrame(data, index=idx)
    df.dropna(inplace=True)

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    for rowy,row in enumerate(dataframe_to_rows(df, index=True, header=False), start=9):
        i = 0
        suma = 0
        if rowy > 9:
            for colx, value in enumerate(row, start=1):
                    collet = get_column_letter(colx)
                    if colx > 2:
                        tot = cols[colx-3][1]
                        cal = calc_nota(int(tot), value)
                        ws.cell(column=colx+i, row=rowy, value=tot) 
                        ws.cell(column=colx+i+1, row=rowy, value=value)
                        ws.cell(column=colx+i+2, row=rowy, value=cal) 
                        suma = suma + value
                        i += 2
                    else: 
                        ws.cell(column=colx, row=rowy, value=value) 
                        ws.cell(column=colx, row=rowy, value=value) 
            suma = round((suma/len(cols)),0)
            pasa = suma
            if pasa > 2:
                form = "SI"
            else:
                form = "NO"
            ws.cell(column=21, row=rowy, value=f'={len(cols)*5}')
            ws.cell(column=22, row=rowy, value=suma)
            ws.cell(column=23, row=rowy, value=form)

    wb.save(f'planillas/Planilla_{grado}.xlsx')


def calc_nota(pt, pl):
    umbral = round(pt*0.7)
    aux = pt - umbral + 1
    divs = trunc(aux/4)
    resto = aux%4
    lim_2 = umbral + divs -  1
    lim_3 = lim_2 + divs + 1
    lim_4 = lim_3 + divs  + 1

    if resto == 1:
        lim_3 += 1
    elif resto == 2:
        lim_3 += 1
        lim_4 += 1
    elif resto == 3:
        lim_2 += 1
        lim_3 += 1
        lim_4 += 1

    if pl < umbral:
        return 1
    else:
        if pl >= umbral and pl <= lim_2:
            return 2
        elif pl > lim_2 and pl <= lim_3:
            return 3
        elif pl > lim_3 and pl <= lim_4:
            return 4
        return 5